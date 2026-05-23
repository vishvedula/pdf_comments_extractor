import io
import os
import re
from datetime import datetime
from typing import Optional

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from groq import Groq
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from pypdf import PdfReader
from pypdf.generic import IndirectObject

load_dotenv()

GROQ_MODEL = os.environ.get("GROQ_MODEL", "llama-3.3-70b-versatile")


def parse_pdf_date(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    m = re.match(r"D?:?(\d{4})(\d{2})?(\d{2})?(\d{2})?(\d{2})?(\d{2})?", raw)
    if not m:
        return raw
    parts = [p for p in m.groups() if p]
    if not parts:
        return raw
    fmt = ["%Y", "%m", "%d", "%H", "%M", "%S"][: len(parts)]
    try:
        dt = datetime.strptime("".join(parts), "".join(fmt))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return raw


def _text(obj) -> str:
    if obj is None:
        return ""
    return str(obj).strip()


def extract_comments(data: bytes, filename: str) -> list[dict]:
    reader = PdfReader(io.BytesIO(data))

    raw_annots = {}
    for page_index, page in enumerate(reader.pages, start=1):
        annots = page.get("/Annots")
        if not annots:
            continue
        for ref in annots:
            if isinstance(ref, IndirectObject):
                key = ref.idnum
                obj = ref.get_object()
            else:
                key = id(ref)
                obj = ref
            if obj is None:
                continue

            subtype = _text(obj.get("/Subtype")).lstrip("/")
            if subtype.lower() == "popup":
                continue

            irt = obj.get("/IRT")
            irt_key = irt.idnum if isinstance(irt, IndirectObject) else None

            raw_annots[key] = {
                "key": key,
                "page": page_index,
                "type": subtype,
                "author": _text(obj.get("/T")),
                "content": _text(obj.get("/Contents")),
                "subject": _text(obj.get("/Subj")),
                "date": parse_pdf_date(
                    _text(obj.get("/M")) or _text(obj.get("/CreationDate"))
                ),
                "irt": irt_key,
            }

    children = {}
    for k, a in raw_annots.items():
        if a["irt"] and a["irt"] in raw_annots:
            children.setdefault(a["irt"], []).append(a)

    rows = []
    seq = 1
    for k, a in raw_annots.items():
        if a["irt"] and a["irt"] in raw_annots:
            continue
        if not a["content"] and not children.get(k):
            continue

        parts = []
        if a["content"]:
            parts.append(a["content"])
        for reply in sorted(children.get(k, []), key=lambda r: r["date"] or ""):
            author = reply["author"] or "Unknown"
            date = reply["date"] or "n/a"
            parts.append(f"↳ Reply by {author} ({date}): {reply['content']}")

        rows.append(
            {
                "S No.": seq,
                "File Name Reviewed": filename,
                "Reference Tab": f"Page {a['page']}",
                "Reviewer": a["author"] or "Unknown",
                "Date Received": a["date"] or "",
                "Review Comments": "\n".join(parts),
            }
        )
        seq += 1
    return rows


def resolve_keyword_columns(kw_df: pd.DataFrame) -> tuple[str, str]:
    norm = {str(c).strip().lower(): c for c in kw_df.columns}
    code_col = norm.get("error code") or norm.get("errorcode") or norm.get("code")
    kw_col = norm.get("keywords") or norm.get("keyword")
    if not code_col or not kw_col:
        raise ValueError(
            f"Expected columns 'Error Code' and 'Keywords'. "
            f"Found: {list(kw_df.columns)}"
        )
    return code_col, kw_col


def build_keyword_reference(
    kw_df: pd.DataFrame, code_col: str, kw_col: str
) -> tuple[str, set[str]]:
    grouped: dict[str, list[str]] = {}
    for _, row in kw_df.iterrows():
        code = str(row[code_col]).strip()
        keyword = str(row[kw_col]).strip()
        if not code or code.lower() == "nan":
            continue
        if not keyword or keyword.lower() == "nan":
            continue
        grouped.setdefault(code, []).append(keyword)
    lines = [f"- {code}: {', '.join(kws)}" for code, kws in grouped.items()]
    return "\n".join(lines), set(grouped)


def classify_comment(
    client: Groq, comment: str, keyword_ref: str, valid_codes: set[str]
) -> str:
    code_list = ", ".join(sorted(valid_codes))
    system = (
        "You classify engineering review comments by assigning each one to an "
        "error code from a provided reference list. Each code has associated "
        "keywords that illustrate the THEME of issues it covers — treat them "
        "as examples, not a strict word-match list.\n\n"
        "Reason semantically about the comment's intent:\n"
        "- A code with keywords like 'Font' covers any presentation, "
        "formatting, wording, spelling, renaming, capitalization, grammar, or "
        "layout issue.\n"
        "- A code with keywords like 'Limit', 'Error', 'Fatigue Life' covers "
        "any engineering, calculation, value, method, or technical-substance "
        "issue.\n"
        "- A code with keywords like 'Missing', 'Not available' covers "
        "anything absent, omitted, or that needs to be added.\n"
        "- Other codes follow their own keyword theme.\n\n"
        "Pick the code whose theme best matches the comment's MEANING, even "
        "if no keyword appears literally. Reserve 'N/A' only for comments "
        "that are empty, off-topic, or genuinely impossible to classify — "
        "never default to N/A out of caution. When in doubt, pick the closest "
        "reasonable code.\n\n"
        "Return only the error code, nothing else."
    )
    user = (
        f"Allowed error codes: {code_list}\n\n"
        f"Reference (error code: example keywords):\n{keyword_ref}\n\n"
        f"Comment:\n{comment}\n\n"
        f"Error Code:"
    )
    resp = client.chat.completions.create(
        model=GROQ_MODEL,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        temperature=0,
        max_tokens=20,
    )
    result = (resp.choices[0].message.content or "").strip()
    result = result.split()[0].strip(".,:;\"'") if result else "N/A"
    if result in valid_codes:
        return result
    lower_map = {c.lower(): c for c in valid_codes}
    if result.lower() in lower_map:
        return lower_map[result.lower()]
    return "N/A"


st.set_page_config(
    page_title="PDF Comments Intelligence",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
.stApp { background: #1e0f3a; color: #f0e6ff; }
#MainMenu, footer, header { visibility: hidden; }

section[data-testid="stSidebar"] {
    background: #180a30;
    border-right: 1px solid #3d2a5c;
}
section[data-testid="stSidebar"] * { color: #e0d4ff; }

h1, h2, h3, h4 { color: #ffffff; }

.brand-title {
    display: flex; align-items: center; gap: 0.75rem;
    font-size: 2.6rem; font-weight: 800; color: #ffffff;
    margin: 0.5rem 0 0;
    font-family: ui-sans-serif, system-ui, sans-serif;
}
.brand-title .bolt { color: #fb923c; font-size: 2.2rem; }
.brand-stack { color: #b8a3d9; margin: 0.25rem 0 1.5rem; font-size: 1rem; }

.section-label {
    color: #c084fc; font-size: 0.85rem; font-weight: 700;
    letter-spacing: 0.18em; text-transform: uppercase;
    margin: 1.5rem 0 0.75rem;
    border-bottom: 1px solid #3d2a5c; padding-bottom: 0.5rem;
}

.pipeline-row {
    display: flex; align-items: center; gap: 0.5rem;
    overflow-x: auto; padding: 0.5rem 0 1rem;
}
.pipeline-card {
    background: #2a1947; border: 1px solid #3d2a5c; border-radius: 14px;
    padding: 1.25rem 0.75rem; text-align: center;
    min-width: 140px; flex: 1;
}
.pipeline-card .icon { font-size: 1.8rem; }
.pipeline-card .name { color: #ffffff; font-weight: 700; margin-top: 0.5rem; font-size: 0.95rem; }
.pipeline-card .desc { color: #b8a3d9; font-size: 0.75rem; margin-top: 0.25rem; }
.pipeline-arrow {
    color: #a78bfa; font-family: ui-monospace, monospace;
    font-weight: 700; font-size: 0.9rem; letter-spacing: -1px;
}

div[data-testid="stFileUploader"] section {
    background: #2a1947 !important;
    border: 1px dashed #c084fc !important;
    border-radius: 12px !important;
}
div[data-testid="stFileUploader"] label,
div[data-testid="stWidgetLabel"] label,
div[data-testid="stWidgetLabel"] p {
    color: #ffffff !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
}
div[data-testid="stFileUploaderDropzoneInstructions"],
div[data-testid="stFileUploaderDropzoneInstructions"] span,
div[data-testid="stFileUploaderDropzoneInstructions"] small,
div[data-testid="stFileUploaderDropzoneInstructions"] div {
    color: #f0e6ff !important;
}
div[data-testid="stFileUploader"] button {
    background: linear-gradient(135deg, #fb923c, #f97316) !important;
    color: #1e0f3a !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 8px !important;
}
div[data-testid="stFileUploader"] button:hover { filter: brightness(1.1); }

.stButton > button,
div[data-testid="stDownloadButton"] > button,
.stDownloadButton > button {
    background: linear-gradient(135deg, #fb923c, #f97316) !important;
    color: #1e0f3a !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.6rem 1.5rem !important;
}
.stButton > button:hover,
div[data-testid="stDownloadButton"] > button:hover { filter: brightness(1.1); }

div[data-testid="stMetric"] {
    background: #2a1947; border: 1px solid #3d2a5c;
    border-radius: 12px; padding: 1rem;
}
div[data-testid="stMetricLabel"], div[data-testid="stMetricLabel"] p {
    color: #c084fc !important;
    font-weight: 600 !important;
}
div[data-testid="stMetricValue"] {
    color: #ffffff !important;
}

div[data-testid="stAlert"] {
    background: #2a1947 !important;
    border: 1px solid #3d2a5c !important;
    color: #f0e6ff !important;
}
div[data-testid="stAlert"] * { color: #f0e6ff !important; }

div[data-testid="stCaptionContainer"], .stCaption {
    color: #b8a3d9 !important;
}

div[data-baseweb="tab-list"] {
    background: transparent; border-bottom: 1px solid #3d2a5c;
    gap: 0.5rem;
}
button[data-baseweb="tab"] {
    background: #2a1947 !important; color: #e0d4ff !important;
    border-radius: 10px 10px 0 0 !important;
    padding: 0.6rem 1.5rem !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    background: #3d2a5c !important; color: #ffffff !important;
    border-bottom: 2px solid #c084fc !important;
}

.sb-brand {
    display: flex; align-items: center; gap: 0.5rem;
    font-size: 1.4rem; font-weight: 800; color: #ffffff;
    margin-top: 0.5rem;
}
.sb-brand .bolt { color: #fb923c; }
.sb-subtitle {
    color: #b8a3d9; font-size: 0.7rem; letter-spacing: 0.2em;
    padding-bottom: 1rem; margin-bottom: 0.5rem;
    border-bottom: 1px solid #3d2a5c;
}
.sb-section {
    color: #b8a3d9; font-size: 0.7rem; letter-spacing: 0.2em;
    margin: 1.5rem 0 0.5rem;
}
.sb-item {
    padding: 0.45rem 0; color: #e0d4ff; font-size: 0.95rem;
}
.sb-footer {
    margin-top: 2.5rem; padding-top: 1rem;
    border-top: 1px solid #3d2a5c; color: #b8a3d9; font-size: 0.85rem;
}
.sb-footer .label {
    font-size: 0.7rem; letter-spacing: 0.2em; margin-bottom: 0.4rem;
}
</style>
""",
    unsafe_allow_html=True,
)

with st.sidebar:
    st.markdown(
        """
<div class="sb-brand"><span class="bolt">⚡</span><span>PDF COMMENTS</span></div>
<div class="sb-subtitle">AUTONOMOUS EXTRACTOR</div>

<div class="sb-section">NAVIGATION</div>
<div class="sb-item">📥 &nbsp; Upload</div>
<div class="sb-item">📝 &nbsp; Extract</div>
<div class="sb-item">🧠 &nbsp; Classify</div>
<div class="sb-item">📊 &nbsp; Visualize</div>
<div class="sb-item">⬇️ &nbsp; Download</div>

<div class="sb-footer">
    <div class="label">POWERED BY</div>
    pypdf · LLaMA 3.3<br>
    Groq · Streamlit
</div>
""",
        unsafe_allow_html=True,
    )

st.markdown(
    """
<h1 class="brand-title"><span class="bolt">⚡</span> PDF Comments Intelligence</h1>
<div class="brand-stack">pypdf · LLaMA 3.3 · Groq · Multi-Stage Pipeline</div>
""",
    unsafe_allow_html=True,
)

st.markdown('<div class="section-label">PIPELINE</div>', unsafe_allow_html=True)
st.markdown(
    """
<div class="pipeline-row">
    <div class="pipeline-card">
        <div class="icon">📄</div>
        <div class="name">Upload</div>
        <div class="desc">PDF &amp; keywords</div>
    </div>
    <div class="pipeline-arrow">===&gt;</div>
    <div class="pipeline-card">
        <div class="icon">📝</div>
        <div class="name">Extract</div>
        <div class="desc">Annotations &amp; replies</div>
    </div>
    <div class="pipeline-arrow">===&gt;</div>
    <div class="pipeline-card">
        <div class="icon">🧠</div>
        <div class="name">Classify</div>
        <div class="desc">LLaMA error codes</div>
    </div>
    <div class="pipeline-arrow">===&gt;</div>
    <div class="pipeline-card">
        <div class="icon">📊</div>
        <div class="name">Visualize</div>
        <div class="desc">Charts &amp; summary</div>
    </div>
    <div class="pipeline-arrow">===&gt;</div>
    <div class="pipeline-card">
        <div class="icon">⬇️</div>
        <div class="name">Export</div>
        <div class="desc">CSV &amp; Excel</div>
    </div>
</div>
""",
    unsafe_allow_html=True,
)

st.markdown('<div class="section-label">QUICK RUN</div>', unsafe_allow_html=True)
col_a, col_b = st.columns(2)
uploaded = col_a.file_uploader("PDF file", type=["pdf"])
kw_uploaded = col_b.file_uploader(
    "Keywords sheet (optional, .xlsx)", type=["xlsx"]
)

if uploaded is not None and st.button("Extract Comments", type="primary"):
    with st.spinner("Extracting comments..."):
        try:
            rows = extract_comments(uploaded.getvalue(), uploaded.name)
        except Exception as e:
            st.error(f"Could not process PDF: {e}")
            st.stop()

    if not rows:
        st.warning("No comments were found in this PDF.")
        st.stop()

    df = pd.DataFrame(rows)

    if kw_uploaded is not None:
        api_key = os.environ.get("GROQ_API_KEY")
        if not api_key:
            st.warning(
                "Keywords file uploaded but GROQ_API_KEY is not set in .env — "
                "skipping error-code classification."
            )
        else:
            try:
                kw_df = pd.read_excel(kw_uploaded)
                try:
                    code_col, kw_col = resolve_keyword_columns(kw_df)
                except ValueError as ve:
                    st.error(str(ve))
                    st.write("First few rows of the uploaded sheet:")
                    st.dataframe(kw_df.head())
                    st.stop()
                keyword_ref, valid_codes = build_keyword_reference(
                    kw_df, code_col, kw_col
                )
                client = Groq(api_key=api_key)

                codes = []
                progress = st.progress(0.0, text="Classifying comments...")
                for i, comment in enumerate(df["Review Comments"], start=1):
                    try:
                        codes.append(
                            classify_comment(client, comment, keyword_ref, valid_codes)
                        )
                    except Exception as e:
                        st.warning(f"Row {i}: classification failed ({e})")
                        codes.append("N/A")
                    progress.progress(i / len(df), text=f"Classified {i}/{len(df)}")
                progress.empty()
                df["Error Code"] = codes
            except Exception as e:
                st.error(f"Classification step failed: {e}")
                st.stop()

    st.markdown(
        '<div class="section-label">RESULTS</div>', unsafe_allow_html=True
    )
    m1, m2, m3 = st.columns(3)
    m1.metric("Comment Threads", len(df))
    m2.metric(
        "Reviewers",
        df["Reviewer"].nunique() if "Reviewer" in df.columns else 0,
    )
    m3.metric(
        "Error Codes",
        df["Error Code"].nunique() if "Error Code" in df.columns else 0,
    )

    base = uploaded.name.rsplit(".", 1)[0]

    summary_df = None
    if "Error Code" in df.columns:
        summary_df = (
            df["Error Code"]
            .value_counts()
            .rename_axis("Error Code")
            .reset_index(name="Count")
        )

    tab_comments, tab_summary = st.tabs(["Comments", "Summary"])
    with tab_comments:
        st.dataframe(df, use_container_width=True, hide_index=True)
    with tab_summary:
        if summary_df is not None and not summary_df.empty:
            st.subheader(f"Error Code Distribution — {base}")
            st.bar_chart(summary_df.set_index("Error Code")["Count"])
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
        else:
            st.info(
                "Upload a keywords sheet and configure GROQ_API_KEY to "
                "enable error-code classification and the summary chart."
            )

    csv_bytes = df.to_csv(index=False).encode("utf-8")

    xlsx_buf = io.BytesIO()
    with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Comments")
        comments_ws = writer.sheets["Comments"]
        header_fill = PatternFill(
            start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"
        )
        header_font = Font(bold=True)
        centered = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        for cell in comments_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered
        for row in comments_ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = centered
        for column_cells in comments_ws.columns:
            comments_ws.column_dimensions[column_cells[0].column_letter].width = 40

        if summary_df is not None and not summary_df.empty:
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
            ws = writer.sheets["Summary"]
            n_rows = len(summary_df)
            chart = BarChart()
            chart.type = "col"
            chart.title = base
            chart.x_axis.title = "Error Code"
            chart.y_axis.title = "Count"
            chart.legend = None
            data = Reference(ws, min_col=2, min_row=1, max_row=n_rows + 1, max_col=2)
            cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "D2")
    xlsx_bytes = xlsx_buf.getvalue()
    c1, c2 = st.columns(2)
    c1.download_button(
        "Download CSV",
        csv_bytes,
        file_name=f"{base}_comments.csv",
        mime="text/csv",
    )
    c2.download_button(
        "Download Excel",
        xlsx_bytes,
        file_name=f"{base}_comments.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
